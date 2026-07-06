"""File upload/save, kind detection, dataset discovery and row readers (csv/xlsx/jsonl/sqlite)."""
from __future__ import annotations

import csv
import json
import os
import sqlite3
import uuid
from collections.abc import Iterable
from typing import Any

from app.core.config import settings
from app.services.structured_common import (
    MAX_PROFILE_ROWS,
    SUPPORTED_FILE_EXTENSIONS,
    LoadedTable,
    cell_to_json,
    infer_runtime_type,
    infer_shape_kind,
    quote_sqlite_ident,
    safe_filename,
    stringify_cell,
)
from app.services.structured_store import StructuredStore, get_structured_store


def get_upload_dir(owner_id: str) -> str:
    root = os.path.join(settings.DATA_DIR, "structured_uploads", safe_filename(owner_id))
    os.makedirs(root, exist_ok=True)
    return root


def extension_kind(filename: str) -> str:
    ext = os.path.splitext(filename.lower())[1]
    kind = SUPPORTED_FILE_EXTENSIONS.get(ext)
    if not kind:
        allowed = ", ".join(sorted(SUPPORTED_FILE_EXTENSIONS))
        raise ValueError(f"unsupported structured file type: {ext or filename}; allowed: {allowed}")
    return kind


def save_structured_upload(*, owner_id: str, filename: str, content: bytes) -> tuple[str, str]:
    kind = extension_kind(filename)
    stored = f"{uuid.uuid4().hex}_{safe_filename(filename)}"
    path = os.path.join(get_upload_dir(owner_id), stored)
    with open(path, "wb") as fh:
        fh.write(content)
    return path, kind


def save_structured_upload_stream(*, owner_id: str, filename: str, fileobj) -> tuple[str, str]:
    """流式落盘 + 大小上限（大表 xlsx/csv 整读进内存曾是 OOM 缺口）。"""
    kind = extension_kind(filename)
    # 纯中文文件名坑：safe_filename("项目概算.xlsx") → "_.xlsx" → strip("._")
    # 连点一起剥掉 → 存盘名无扩展名 → openpyxl 按空格式抛 InvalidFileException
    # （线上 500 真栈）。显式补回已通过校验的真实扩展名。
    ext = os.path.splitext(filename.lower())[1]
    stored = f"{uuid.uuid4().hex}_{safe_filename(filename)}"
    if ext and not stored.lower().endswith(ext):
        stored = f"{stored}{ext}"
    path = os.path.join(get_upload_dir(owner_id), stored)
    max_bytes = int(getattr(settings, "STRUCTURED_MAX_FILE_SIZE", 200 * 1024**2))
    copy_stream_with_limit(fileobj, path, max_bytes)
    return path, kind


def register_file_source(
    *,
    owner_id: str,
    filename: str,
    file_path: str,
    kind: str,
    store: StructuredStore | None = None,
) -> dict[str, Any]:
    store = store or get_structured_store()
    source = store.create_source(
        owner_id=owner_id,
        source_type="file",
        kind=kind,
        name=filename,
        file_path=file_path,
        metadata={"original_filename": filename, "file_size": os.path.getsize(file_path)},
    )
    try:
        discovered = discover_file_datasets(source, owner_id=owner_id)
    except ValueError:
        raise
    except Exception as exc:
        # 解析器异常（如 openpyxl InvalidFileException）转干净 400 而非裸 500
        raise ValueError(f"无法解析文件「{filename}」：{exc}") from exc
    for dataset in discovered:
        store.upsert_dataset(owner_id=owner_id, **dataset)
    datasets = store.list_datasets(owner_id=owner_id, source_id=source["id"])
    return {"source": source, "datasets": datasets}


def discover_file_datasets(source: dict[str, Any], *, owner_id: str) -> list[dict[str, Any]]:
    del owner_id
    kind = str(source["kind"])
    path = str(source.get("file_path") or "")
    source_id = str(source["id"])
    if kind == "csv":
        table = read_csv(path, limit=MAX_PROFILE_ROWS)
        return [
            dataset_payload(
                source_id=source_id,
                source_kind=kind,
                name=source["name"],
                table=table,
                metadata={"path": path, "encoding": detect_csv_encoding(path)},
            )
        ]
    if kind == "jsonl":
        table = read_jsonl(path, limit=MAX_PROFILE_ROWS)
        return [
            dataset_payload(
                source_id=source_id,
                source_kind=kind,
                name=source["name"],
                table=table,
                metadata={"path": path},
            )
        ]
    if kind == "xlsx":
        return discover_xlsx_datasets(path, source_id=source_id, source_name=source["name"])
    if kind == "sqlite":
        return discover_sqlite_datasets(path, source_id=source_id, source_kind=kind)
    raise ValueError(f"unsupported file kind: {kind}")


def dataset_payload(
    *,
    source_id: str | None,
    source_kind: str,
    name: str,
    table: LoadedTable,
    metadata: dict[str, Any] | None = None,
    dataset_type: str = "file_table",
    connection_id: str | None = None,
    schema_name: str | None = None,
    table_name: str | None = None,
) -> dict[str, Any]:
    schema = [{"name": col, "data_type": infer_runtime_type([row.get(col) for row in table.rows])} for col in table.columns]
    return {
        "source_id": source_id,
        "connection_id": connection_id,
        "name": name,
        "dataset_type": dataset_type,
        "source_kind": source_kind,
        "shape_kind": infer_shape_kind(table.columns, table.rows),
        "schema_name": schema_name,
        "table_name": table_name,
        "row_count_estimate": table.row_count_estimate,
        "column_count": len(table.columns),
        "schema": schema,
        "metadata": metadata or {},
    }


def detect_csv_encoding(path: str) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk", "latin-1"):
        try:
            with open(path, encoding=encoding, newline="") as fh:
                fh.read(4096)
            return encoding
        except UnicodeDecodeError:
            continue
    return "utf-8"


def read_csv(path: str, *, limit: int | None = None) -> LoadedTable:
    encoding = detect_csv_encoding(path)
    with open(path, encoding=encoding, newline="") as fh:
        sample = fh.read(8192)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(fh, dialect=dialect)
        columns = normalize_columns(reader.fieldnames or [])
        rows: list[dict[str, Any]] = []
        total = 0
        for raw in reader:
            total += 1
            if limit is None or len(rows) < limit:
                rows.append({columns[i]: value for i, value in enumerate(raw.values()) if i < len(columns)})
    return LoadedTable(columns=columns, rows=rows, row_count_estimate=total)


def read_jsonl(path: str, *, limit: int | None = None) -> LoadedTable:
    rows: list[dict[str, Any]] = []
    columns: list[str] = []
    seen: set[str] = set()
    total = 0
    with open(path, encoding="utf-8-sig") as fh:
        for line in fh:
            if not line.strip():
                continue
            total += 1
            obj = json.loads(line)
            if not isinstance(obj, dict):
                obj = {"value": obj}
            for key in obj:
                if key not in seen:
                    seen.add(key)
                    columns.append(str(key))
            if limit is None or len(rows) < limit:
                rows.append({str(key): value for key, value in obj.items()})
    return LoadedTable(columns=normalize_columns(columns), rows=rows, row_count_estimate=total)


def discover_xlsx_datasets(path: str, *, source_id: str, source_name: str) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    datasets: list[dict[str, Any]] = []
    try:
        for sheet_name in wb.sheetnames:
            table = read_xlsx_sheet(path, sheet_name=sheet_name, limit=MAX_PROFILE_ROWS)
            datasets.append(
                dataset_payload(
                    source_id=source_id,
                    source_kind="xlsx",
                    name=f"{source_name} / {sheet_name}",
                    table=table,
                    metadata={"path": path, "sheet_name": sheet_name},
                    dataset_type="sheet",
                )
            )
    finally:
        wb.close()
    return datasets


def read_xlsx_sheet(path: str, *, sheet_name: str, limit: int | None = None) -> LoadedTable:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        iterator = ws.iter_rows(values_only=True)
        header = next(iterator, None)
        columns = normalize_columns([stringify_cell(value) for value in (header or [])])
        if not columns:
            columns = [f"column_{i + 1}" for i in range(ws.max_column or 0)]
        rows: list[dict[str, Any]] = []
        total = 0
        for values in iterator:
            if not values or all(value is None for value in values):
                continue
            total += 1
            if limit is None or len(rows) < limit:
                rows.append({columns[i]: cell_to_json(value) for i, value in enumerate(values[: len(columns)])})
        return LoadedTable(columns=columns, rows=rows, row_count_estimate=total)
    finally:
        wb.close()


def discover_sqlite_datasets(path: str, *, source_id: str | None, source_kind: str) -> list[dict[str, Any]]:
    datasets: list[dict[str, Any]] = []
    with sqlite3.connect(path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT name, type
            FROM sqlite_master
            WHERE type IN ('table', 'view') AND name NOT LIKE 'sqlite_%'
            ORDER BY name
            """
        ).fetchall()
        for row in rows:
            table_name = str(row["name"])
            table = read_sqlite_table(path, table_name=table_name, limit=MAX_PROFILE_ROWS)
            datasets.append(
                dataset_payload(
                    source_id=source_id,
                    source_kind=source_kind,
                    name=table_name,
                    table=table,
                    metadata={"path": path},
                    dataset_type="db_view" if row["type"] == "view" else "db_table",
                    table_name=table_name,
                )
            )
    return datasets


def read_sqlite_table(path: str, *, table_name: str, limit: int | None = None) -> LoadedTable:
    with sqlite3.connect(path) as conn:
        conn.row_factory = sqlite3.Row
        columns = [str(row["name"]) for row in conn.execute(f"PRAGMA table_info({quote_sqlite_ident(table_name)})")]
        total = int(conn.execute(f"SELECT COUNT(*) AS c FROM {quote_sqlite_ident(table_name)}").fetchone()["c"])
        sql = f"SELECT * FROM {quote_sqlite_ident(table_name)}"
        params: tuple[Any, ...] = ()
        if limit is not None:
            sql += " LIMIT ?"
            params = (int(limit),)
        rows = [dict(row) for row in conn.execute(sql, params).fetchall()]
    return LoadedTable(columns=columns, rows=rows, row_count_estimate=total)


def normalize_columns(columns: Iterable[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for idx, raw in enumerate(columns):
        name = str(raw or "").strip() or f"column_{idx + 1}"
        count = seen.get(name, 0)
        seen[name] = count + 1
        out.append(name if count == 0 else f"{name}_{count + 1}")
    return out


def load_dataset_rows(
    dataset: dict[str, Any],
    *,
    owner_id: str,
    limit: int | None,
    store: StructuredStore | None = None,
) -> LoadedTable:
    store = store or get_structured_store()
    source_kind = str(dataset.get("source_kind") or "")
    metadata = dataset.get("metadata") or {}
    if dataset.get("source_id"):
        path = str(metadata.get("path") or "")
        if source_kind == "csv":
            return read_csv(path, limit=limit)
        if source_kind == "jsonl":
            return read_jsonl(path, limit=limit)
        if source_kind == "xlsx":
            return read_xlsx_sheet(path, sheet_name=str(metadata.get("sheet_name") or ""), limit=limit)
        if source_kind == "sqlite":
            return read_sqlite_table(path, table_name=str(dataset.get("table_name") or dataset["name"]), limit=limit)
    if dataset.get("connection_id"):
        # 函数内导入：structured_connections 顶层依赖本模块的 reader，避免循环导入
        from app.services.structured_connections import decrypt_credential, read_connection_table

        connection = store.get_connection(str(dataset["connection_id"]), owner_id=owner_id, include_secret=True)
        if not connection:
            raise ValueError("connection not found")
        credential = decrypt_credential(connection.get("credential") or {})
        return read_connection_table(
            connection,
            credential,
            schema_name=dataset.get("schema_name"),
            table_name=str(dataset.get("table_name") or dataset["name"]),
            limit=limit,
        )
    raise ValueError("dataset source not available")


def copy_stream_with_limit(src, dst_path: str, max_bytes: int, chunk_size: int = 1024 * 1024) -> int:
    """分块拷贝上传流并强制大小上限；超限删除半成品并报错（对齐 /files/upload）。"""
    written = 0
    try:
        with open(dst_path, "wb") as dst:
            while True:
                chunk = src.read(chunk_size)
                if not chunk:
                    break
                written += len(chunk)
                if written > max_bytes:
                    raise ValueError(f"文件超过上传上限 {max_bytes // (1024 * 1024)}MB")
                dst.write(chunk)
    except Exception:
        if os.path.exists(dst_path):
            os.remove(dst_path)
        raise
    return written


def iter_dataset_rows(
    dataset: dict[str, Any],
    *,
    owner_id: str,
    store: StructuredStore | None = None,
):
    """流式读数据集：返回 (columns, 行迭代器)，内存 O(1)。

    preview/profile 仍走 load_dataset_rows（有限行）；导出用本函数，
    十万行不再全量进内存。列集合与既有 read_* 的口径一致。
    """
    store = store or get_structured_store()
    source_kind = str(dataset.get("source_kind") or "")
    metadata = dataset.get("metadata") or {}
    if dataset.get("source_id"):
        path = str(metadata.get("path") or "")
        if source_kind == "csv":
            return _iter_csv_rows(path)
        if source_kind == "jsonl":
            return _iter_jsonl_rows(path)
        if source_kind == "xlsx":
            return _iter_xlsx_rows(path, sheet_name=str(metadata.get("sheet_name") or ""))
        if source_kind == "sqlite":
            return _iter_sqlite_rows(path, table_name=str(dataset.get("table_name") or dataset["name"]))
    if dataset.get("connection_id"):
        # 函数内导入：structured_connections 顶层依赖本模块的 reader，避免循环导入
        from app.services.structured_connections import _iter_connection_rows, decrypt_credential

        connection = store.get_connection(str(dataset["connection_id"]), owner_id=owner_id, include_secret=True)
        if not connection:
            raise ValueError("connection not found")
        credential = decrypt_credential(connection.get("credential") or {})
        engine = str(connection.get("engine") or "")
        table_name = str(dataset.get("table_name") or dataset["name"])
        if engine == "sqlite":
            path = str(credential.get("sqlite_path") or credential.get("database") or "")
            return _iter_sqlite_rows(path, table_name=table_name)
        return _iter_connection_rows(
            connection, credential,
            schema_name=dataset.get("schema_name"), table_name=table_name,
        )
    raise ValueError("dataset source not available")


def _iter_csv_rows(path: str):
    encoding = detect_csv_encoding(path)
    with open(path, encoding=encoding, newline="") as fh:
        sample = fh.read(8192)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        fh.seek(0)
        columns = normalize_columns(csv.DictReader(fh, dialect=dialect).fieldnames or [])

    def rows():
        with open(path, encoding=encoding, newline="") as handle:
            reader = csv.DictReader(handle, dialect=dialect)
            for raw in reader:
                yield {columns[i]: value for i, value in enumerate(raw.values()) if i < len(columns)}

    return columns, rows()


def _iter_jsonl_rows(path: str, probe_limit: int = MAX_PROFILE_ROWS):
    columns: list[str] = []
    seen: set[str] = set()
    with open(path, encoding="utf-8-sig") as fh:
        for index, line in enumerate(fh):
            if index >= probe_limit:
                break
            if not line.strip():
                continue
            obj = json.loads(line)
            if not isinstance(obj, dict):
                obj = {"value": obj}
            for key in obj:
                if key not in seen:
                    seen.add(key)
                    columns.append(str(key))
    normalized = normalize_columns(columns)

    def rows():
        with open(path, encoding="utf-8-sig") as handle:
            for line in handle:
                if not line.strip():
                    continue
                obj = json.loads(line)
                if not isinstance(obj, dict):
                    obj = {"value": obj}
                yield {str(key): value for key, value in obj.items()}

    return normalized, rows()


def _iter_xlsx_rows(path: str, *, sheet_name: str):
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        header = next(ws.iter_rows(values_only=True), None)
        columns = normalize_columns([stringify_cell(value) for value in (header or [])])
        if not columns:
            columns = [f"column_{i + 1}" for i in range(ws.max_column or 0)]
    finally:
        wb.close()

    def rows():
        book = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = book[sheet_name]
            iterator = sheet.iter_rows(values_only=True)
            next(iterator, None)  # skip header
            for values in iterator:
                if not values or all(value is None for value in values):
                    continue
                yield {columns[i]: cell_to_json(value) for i, value in enumerate(values[: len(columns)])}
        finally:
            book.close()

    return columns, rows()


def _iter_sqlite_rows(path: str, *, table_name: str):
    with sqlite3.connect(path) as conn:
        columns = [str(row[1]) for row in conn.execute(f"PRAGMA table_info({quote_sqlite_ident(table_name)})")]

    def rows():
        with sqlite3.connect(path) as conn:
            conn.row_factory = sqlite3.Row
            for row in conn.execute(f"SELECT * FROM {quote_sqlite_ident(table_name)}"):
                yield dict(row)

    return columns, rows()
