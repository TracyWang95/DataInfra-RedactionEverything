"""Structured data service that bypasses OCR and model inference."""
from __future__ import annotations

import base64
import csv
import hashlib
import hmac
import ipaddress
import json
import logging
import os
import re
import sqlite3
import threading
import time
import uuid
import zipfile
from collections import Counter
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import httpx
from cryptography.fernet import Fernet

from app.core.config import settings
from app.models.type_mapping import canonical_type_id, cn_to_id
from app.services.structured_store import StructuredStore, get_structured_store

logger = logging.getLogger(__name__)

SUPPORTED_FILE_EXTENSIONS = {".csv": "csv", ".xlsx": "xlsx", ".jsonl": "jsonl", ".db": "sqlite", ".sqlite": "sqlite"}
MAX_PROFILE_ROWS = 500
MAX_PREVIEW_ROWS = 100
MAX_EXPORT_ROWS = 250_000
_STRUCTURED_EXPORT_LOCK = threading.Lock()


@dataclass(frozen=True)
class LoadedTable:
    columns: list[str]
    rows: list[dict[str, Any]]
    row_count_estimate: int | None = None


_COLUMN_NAME_HINTS: list[tuple[str, str, str, float]] = [
    (r"(phone|mobile|tel|手机号|手机|电话|联系方式|联系电话)", "PHONE", "high", 0.95),
    (r"(email|mail|邮箱|电子邮件)", "EMAIL", "high", 0.95),
    (r"(id.?card|身份证|证件号|证件号码|identity)", "ID_CARD", "critical", 0.95),
    (r"(bank.?card|银行卡号|银行账号|账户|账号)", "BANK_CARD", "critical", 0.88),
    (r"(password|passwd|pwd|密码|口令)", "USERNAME_PASSWORD", "critical", 0.98),
    (r"(token|secret|key|api.?key|密钥|令牌|凭证)", "AUTH_SECRET", "critical", 0.96),
    (r"(company|corp|org|机构|公司|单位|供应商|客户|企业)", "ORG", "high", 0.86),
    (
        r"(^name$|full.?name|customer.?name|user.?name|receiver.?name|contact.?name|person.?name|"
        r"employee.?name|staff.?name|agent.?name|account.?name|payer.?name|payee.?name|owner.?name|"
        r"legal.?representative|姓名|联系人|客户名|用户名|收件人|经办人|代理人|负责人|开户名|账户名)",
        "PERSON",
        "high",
        0.72,
    ),
    (r"(address|addr|住址|地址|门牌|地区|省|市|区县)", "ADDRESS", "medium", 0.82),
    (r"(amount|price|money|salary|fee|金额|价格|单价|合计|余额|费用|收入|支出)", "AMOUNT", "medium", 0.82),
    (r"(date|time|created|updated|生日|出生|日期|时间)", "DATE", "medium", 0.72),
    (r"(ip地址|ip$|ip_|ipaddress)", "IP_ADDRESS", "medium", 0.92),
    (r"(mac地址|mac$|mac_)", "MAC_ADDRESS", "medium", 0.9),
    (r"(url|website|site|网址|链接)", "URL_WEBSITE", "medium", 0.88),
    (r"(license|plate|车牌)", "LICENSE_PLATE", "high", 0.84),
    (r"(contract|合同号|订单号|单据号|编号|流水号)", "DOCUMENT_NUMBER", "medium", 0.7),
]

_BUSINESS_DESCRIPTOR_COLUMN_PATTERNS = [
    re.compile(
        r"(^|[_\-\s])("
        r"product|sku|item|goods|commodity|material|device|equipment|model|brand|category|catalog|"
        r"spec|title|subject|project|service|package|plan|version|status|type|tier|memo|note|summary|"
        r"description|content"
        r")([_\-\s]|$)",
        re.I,
    ),
    re.compile(r"(产品|商品|物料|设备|装备|器材|型号|规格|品牌|类目|品类|标题|主题|项目|服务|套餐|方案|版本|状态|类型|等级|备注|摘要|描述|说明|内容)"),
]

_BUSINESS_DESCRIPTOR_BLOCK_TYPES = {
    "PERSON",
    "PHONE",
    "EMAIL",
    "ID_CARD",
    "BANK_CARD",
    "USERNAME_PASSWORD",
    "AUTH_SECRET",
    "ADDRESS",
    "IP_ADDRESS",
    "MAC_ADDRESS",
    "LICENSE_PLATE",
}

_VALUE_PATTERNS: list[tuple[re.Pattern[str], str, str, float]] = [
    (re.compile(r"^1[3-9]\d{9}$"), "PHONE", "high", 0.98),
    (re.compile(r"^[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}$", re.I), "EMAIL", "high", 0.98),
    (re.compile(r"^\d{17}[\dXx]$"), "ID_CARD", "critical", 0.98),
    (re.compile(r"^\d{12,19}$"), "BANK_CARD", "critical", 0.72),
    (re.compile(r"^(?:\d{1,3}\.){3}\d{1,3}$"), "IP_ADDRESS", "medium", 0.94),
    (re.compile(r"^[0-9A-F]{2}(?::[0-9A-F]{2}){5}$", re.I), "MAC_ADDRESS", "medium", 0.94),
    (re.compile(r"^https?://", re.I), "URL_WEBSITE", "medium", 0.92),
    (re.compile(r"^\d{4}(?:[-/\u5e74]\d{1,2}(?:[-/\u6708]\d{1,2}\u65e5?)?)?$"), "DATE", "medium", 0.78),
    (re.compile(r"^-?[\u00a5\uffe5$]?\d{1,3}(?:,\d{3})*(?:\.\d+)?$|^-?[\u00a5\uffe5$]?\d+(?:\.\d+)?$"), "AMOUNT", "medium", 0.55),
]

# Type-inference acceptance thresholds + shape heuristics (names for literals).
_WIDE_TABLE_COLUMN_THRESHOLD = 80
_CUSTOM_TYPE_CONFIDENCE_MIN = 0.55
_DEFAULT_TYPE_CONFIDENCE_MIN = 0.78
_DEFAULT_DATASET_DISCOVERY_LIMIT = 500

_PII_DEFAULT_MASK_TYPES = {
    "PHONE",
    "EMAIL",
    "ID_CARD",
    "BANK_CARD",
    "LICENSE_PLATE",
}
_PII_DEFAULT_HASH_TYPES = {"IP_ADDRESS", "MAC_ADDRESS", "DEVICE_ID"}
_PII_DEFAULT_GENERALIZE_TYPES = {"ADDRESS"}
_PII_DEFAULT_TOKENIZE_TYPES = {"PERSON"}
_SECURITY_DEFAULT_SUPPRESS_TYPES = {"USERNAME_PASSWORD", "AUTH_SECRET"}

_STRUCTURED_HAS_NER_TYPES = [
    "\u59d3\u540d",
    "\u7535\u8bdd",
    "\u90ae\u7bb1",
    "\u8eab\u4efd\u8bc1\u53f7",
    "\u94f6\u884c\u5361\u53f7",
    "\u94f6\u884c\u8d26\u53f7",
    "\u5730\u5740",
    "\u8f66\u724c",
    "IP\u5730\u5740",
    "MAC\u5730\u5740",
    "\u767b\u5f55\u8d26\u53f7",
    "\u5bc6\u7801",
    "\u7f51\u5740\u94fe\u63a5",
    "\u8bbe\u5907\u53f7",
    "\u62a4\u7167\u53f7",
    "\u793e\u4fdd\u53f7",
]
_STRUCTURED_SEMANTIC_TYPE_RISK = {
    "PERSON": "high",
    "PHONE": "high",
    "EMAIL": "high",
    "ID_CARD": "critical",
    "BANK_CARD": "critical",
    "BANK_ACCOUNT": "critical",
    "ADDRESS": "medium",
    "LICENSE_PLATE": "high",
    "IP_ADDRESS": "medium",
    "MAC_ADDRESS": "medium",
    "DEVICE_ID": "medium",
    "USERNAME_PASSWORD": "critical",
    "AUTH_SECRET": "critical",
    "PASSPORT": "critical",
    "SOCIAL_SECURITY": "critical",
    "URL_WEBSITE": "medium",
}
_STRUCTURED_DIRECT_VALUE_TYPES = {
    "PHONE",
    "EMAIL",
    "ID_CARD",
    "BANK_CARD",
    "BANK_ACCOUNT",
    "IP_ADDRESS",
    "MAC_ADDRESS",
    "LICENSE_PLATE",
    "URL_WEBSITE",
}
_SEMANTIC_READY_CACHE: tuple[float, bool] = (0.0, False)
_SEMANTIC_READY_TTL_SEC = 15.0


def utc_iso() -> str:
    return datetime.now(UTC).isoformat()


def get_upload_dir(owner_id: str) -> str:
    root = os.path.join(settings.DATA_DIR, "structured_uploads", safe_filename(owner_id))
    os.makedirs(root, exist_ok=True)
    return root


def safe_filename(name: str) -> str:
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", str(name or "data")).strip("._")
    return stem[:120] or "data"


def export_base_filename(name: str) -> str:
    safe = safe_filename(name)
    suffix = Path(safe).suffix.lower()
    if suffix in SUPPORTED_FILE_EXTENSIONS:
        safe = safe[: -len(suffix)].strip("._")
    return safe or "data"


def unique_export_filename(
    filename: str,
    *,
    dataset_id: str,
    owner_id: str,
    job_id: str,
    out_dir: str,
    store: StructuredStore,
) -> str:
    """Keep job export names readable while preventing collisions in the delivery package."""
    existing_names = {
        str(export.get("filename") or "").lower()
        for export in store.list_exports(owner_id=owner_id, job_id=job_id)
        if export.get("filename")
    }
    try:
        existing_names.update(item.lower() for item in os.listdir(out_dir))
    except FileNotFoundError:
        pass

    if filename.lower() not in existing_names and not os.path.exists(os.path.join(out_dir, filename)):
        return filename

    stem, ext = os.path.splitext(filename)
    short_id = safe_filename(dataset_id).replace("-", "")[:8] or uuid.uuid4().hex[:8]
    for index in range(2, 1000):
        candidate = f"{stem}-{index}-{short_id}{ext}"
        if candidate.lower() not in existing_names and not os.path.exists(os.path.join(out_dir, candidate)):
            return candidate
    raise ValueError("unable to allocate a unique structured export filename")


def extension_kind(filename: str) -> str:
    ext = os.path.splitext(filename.lower())[1]
    kind = SUPPORTED_FILE_EXTENSIONS.get(ext)
    if not kind:
        allowed = ", ".join(sorted(SUPPORTED_FILE_EXTENSIONS))
        raise ValueError(f"unsupported structured file type: {ext or filename}; allowed: {allowed}")
    return kind


def credential_key_path() -> str:
    return os.path.join(settings.DATA_DIR, "structured_credentials.key")


def _fernet() -> Fernet:
    path = credential_key_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if not os.path.exists(path):
        with open(path, "wb") as fh:
            fh.write(Fernet.generate_key())
        try:
            os.chmod(path, 0o600)
        except OSError:
            pass
    with open(path, "rb") as fh:
        return Fernet(fh.read().strip())


def encrypt_credential(payload: dict[str, Any]) -> dict[str, str]:
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return {"encrypted": _fernet().encrypt(raw).decode("ascii")}


def decrypt_credential(payload: dict[str, Any]) -> dict[str, Any]:
    token = payload.get("encrypted") if isinstance(payload, dict) else None
    if not isinstance(token, str) or not token:
        return {}
    raw = _fernet().decrypt(token.encode("ascii")).decode("utf-8")
    data = json.loads(raw)
    return data if isinstance(data, dict) else {}


def save_structured_upload(*, owner_id: str, filename: str, content: bytes) -> tuple[str, str]:
    kind = extension_kind(filename)
    stored = f"{uuid.uuid4().hex}_{safe_filename(filename)}"
    path = os.path.join(get_upload_dir(owner_id), stored)
    with open(path, "wb") as fh:
        fh.write(content)
    return path, kind


def save_structured_upload_stream(*, owner_id: str, filename: str, fileobj) -> tuple[str, str]:
    """流式落盘 + 大小上限（大表 xlsx/csv 整读进内存曾是 OOM 缺口）。"""
    kind = extension_kind(filename)
    stored = f"{uuid.uuid4().hex}_{safe_filename(filename)}"
    path = os.path.join(get_upload_dir(owner_id), stored)
    max_bytes = int(getattr(settings, "STRUCTURED_MAX_FILE_SIZE", 200 * 1024**2))
    copy_stream_with_limit(fileobj, path, max_bytes)
    return path, kind


def register_file_source(
    *,
    owner_id: str,
    filename: str,
    file_path: str,
    kind: str,
    store: StructuredStore | None = None,
) -> dict[str, Any]:
    store = store or get_structured_store()
    source = store.create_source(
        owner_id=owner_id,
        source_type="file",
        kind=kind,
        name=filename,
        file_path=file_path,
        metadata={"original_filename": filename, "file_size": os.path.getsize(file_path)},
    )
    for dataset in discover_file_datasets(source, owner_id=owner_id):
        store.upsert_dataset(owner_id=owner_id, **dataset)
    datasets = store.list_datasets(owner_id=owner_id, source_id=source["id"])
    return {"source": source, "datasets": datasets}


def discover_file_datasets(source: dict[str, Any], *, owner_id: str) -> list[dict[str, Any]]:
    del owner_id
    kind = str(source["kind"])
    path = str(source.get("file_path") or "")
    source_id = str(source["id"])
    if kind == "csv":
        table = read_csv(path, limit=MAX_PROFILE_ROWS)
        return [
            dataset_payload(
                source_id=source_id,
                source_kind=kind,
                name=source["name"],
                table=table,
                metadata={"path": path, "encoding": detect_csv_encoding(path)},
            )
        ]
    if kind == "jsonl":
        table = read_jsonl(path, limit=MAX_PROFILE_ROWS)
        return [
            dataset_payload(
                source_id=source_id,
                source_kind=kind,
                name=source["name"],
                table=table,
                metadata={"path": path},
            )
        ]
    if kind == "xlsx":
        return discover_xlsx_datasets(path, source_id=source_id, source_name=source["name"])
    if kind == "sqlite":
        return discover_sqlite_datasets(path, source_id=source_id, source_kind=kind)
    raise ValueError(f"unsupported file kind: {kind}")


def dataset_payload(
    *,
    source_id: str | None,
    source_kind: str,
    name: str,
    table: LoadedTable,
    metadata: dict[str, Any] | None = None,
    dataset_type: str = "file_table",
    connection_id: str | None = None,
    schema_name: str | None = None,
    table_name: str | None = None,
) -> dict[str, Any]:
    schema = [{"name": col, "data_type": infer_runtime_type([row.get(col) for row in table.rows])} for col in table.columns]
    return {
        "source_id": source_id,
        "connection_id": connection_id,
        "name": name,
        "dataset_type": dataset_type,
        "source_kind": source_kind,
        "shape_kind": infer_shape_kind(table.columns, table.rows),
        "schema_name": schema_name,
        "table_name": table_name,
        "row_count_estimate": table.row_count_estimate,
        "column_count": len(table.columns),
        "schema": schema,
        "metadata": metadata or {},
    }


def annotate_discovered_connection_datasets(
    datasets: Iterable[dict[str, Any]],
    *,
    connection_id: str,
) -> list[dict[str, Any]]:
    now = utc_iso()
    annotated: list[dict[str, Any]] = []
    for dataset in datasets:
        item = dict(dataset)
        item["source_id"] = item.get("source_id") or None
        item["connection_id"] = connection_id
        item.setdefault("created_at", now)
        item.setdefault("id", stable_discovered_dataset_id(connection_id, item))
        annotated.append(item)
    return annotated


def stable_discovered_dataset_id(connection_id: str, dataset: dict[str, Any]) -> str:
    schema = str(dataset.get("schema_name") or "")
    table = str(dataset.get("table_name") or dataset.get("name") or "")
    dataset_type = str(dataset.get("dataset_type") or "")
    key = f"structured-discovery:{connection_id}:{schema}:{table}:{dataset_type}"
    return f"discovered:{uuid.uuid5(uuid.NAMESPACE_URL, key).hex}"


def detect_csv_encoding(path: str) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk", "latin-1"):
        try:
            with open(path, encoding=encoding, newline="") as fh:
                fh.read(4096)
            return encoding
        except UnicodeDecodeError:
            continue
    return "utf-8"


def read_csv(path: str, *, limit: int | None = None) -> LoadedTable:
    encoding = detect_csv_encoding(path)
    with open(path, encoding=encoding, newline="") as fh:
        sample = fh.read(8192)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(fh, dialect=dialect)
        columns = normalize_columns(reader.fieldnames or [])
        rows: list[dict[str, Any]] = []
        total = 0
        for raw in reader:
            total += 1
            if limit is None or len(rows) < limit:
                rows.append({columns[i]: value for i, value in enumerate(raw.values()) if i < len(columns)})
    return LoadedTable(columns=columns, rows=rows, row_count_estimate=total)


def read_jsonl(path: str, *, limit: int | None = None) -> LoadedTable:
    rows: list[dict[str, Any]] = []
    columns: list[str] = []
    seen: set[str] = set()
    total = 0
    with open(path, encoding="utf-8-sig") as fh:
        for line in fh:
            if not line.strip():
                continue
            total += 1
            obj = json.loads(line)
            if not isinstance(obj, dict):
                obj = {"value": obj}
            for key in obj:
                if key not in seen:
                    seen.add(key)
                    columns.append(str(key))
            if limit is None or len(rows) < limit:
                rows.append({str(key): value for key, value in obj.items()})
    return LoadedTable(columns=normalize_columns(columns), rows=rows, row_count_estimate=total)


def discover_xlsx_datasets(path: str, *, source_id: str, source_name: str) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    datasets: list[dict[str, Any]] = []
    try:
        for sheet_name in wb.sheetnames:
            table = read_xlsx_sheet(path, sheet_name=sheet_name, limit=MAX_PROFILE_ROWS)
            datasets.append(
                dataset_payload(
                    source_id=source_id,
                    source_kind="xlsx",
                    name=f"{source_name} / {sheet_name}",
                    table=table,
                    metadata={"path": path, "sheet_name": sheet_name},
                    dataset_type="sheet",
                )
            )
    finally:
        wb.close()
    return datasets


def read_xlsx_sheet(path: str, *, sheet_name: str, limit: int | None = None) -> LoadedTable:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        iterator = ws.iter_rows(values_only=True)
        header = next(iterator, None)
        columns = normalize_columns([stringify_cell(value) for value in (header or [])])
        if not columns:
            columns = [f"column_{i + 1}" for i in range(ws.max_column or 0)]
        rows: list[dict[str, Any]] = []
        total = 0
        for values in iterator:
            if not values or all(value is None for value in values):
                continue
            total += 1
            if limit is None or len(rows) < limit:
                rows.append({columns[i]: cell_to_json(value) for i, value in enumerate(values[: len(columns)])})
        return LoadedTable(columns=columns, rows=rows, row_count_estimate=total)
    finally:
        wb.close()


def discover_sqlite_datasets(path: str, *, source_id: str | None, source_kind: str) -> list[dict[str, Any]]:
    datasets: list[dict[str, Any]] = []
    with sqlite3.connect(path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT name, type
            FROM sqlite_master
            WHERE type IN ('table', 'view') AND name NOT LIKE 'sqlite_%'
            ORDER BY name
            """
        ).fetchall()
        for row in rows:
            table_name = str(row["name"])
            table = read_sqlite_table(path, table_name=table_name, limit=MAX_PROFILE_ROWS)
            datasets.append(
                dataset_payload(
                    source_id=source_id,
                    source_kind=source_kind,
                    name=table_name,
                    table=table,
                    metadata={"path": path},
                    dataset_type="db_view" if row["type"] == "view" else "db_table",
                    table_name=table_name,
                )
            )
    return datasets


def read_sqlite_table(path: str, *, table_name: str, limit: int | None = None) -> LoadedTable:
    with sqlite3.connect(path) as conn:
        conn.row_factory = sqlite3.Row
        columns = [str(row["name"]) for row in conn.execute(f"PRAGMA table_info({quote_sqlite_ident(table_name)})")]
        total = int(conn.execute(f"SELECT COUNT(*) AS c FROM {quote_sqlite_ident(table_name)}").fetchone()["c"])
        sql = f"SELECT * FROM {quote_sqlite_ident(table_name)}"
        params: tuple[Any, ...] = ()
        if limit is not None:
            sql += " LIMIT ?"
            params = (int(limit),)
        rows = [dict(row) for row in conn.execute(sql, params).fetchall()]
    return LoadedTable(columns=columns, rows=rows, row_count_estimate=total)


def quote_sqlite_ident(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def normalize_columns(columns: Iterable[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for idx, raw in enumerate(columns):
        name = str(raw or "").strip() or f"column_{idx + 1}"
        count = seen.get(name, 0)
        seen[name] = count + 1
        out.append(name if count == 0 else f"{name}_{count + 1}")
    return out


def stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def cell_to_json(value: Any) -> Any:
    if value is None or isinstance(value, str | int | float | bool):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return str(value)


def infer_runtime_type(values: Iterable[Any]) -> str:
    sample = [value for value in values if value not in (None, "")]
    if not sample:
        return "string"
    if all(isinstance(value, bool) for value in sample):
        return "boolean"
    if all(is_number_like(value) for value in sample):
        return "number"
    return "string"


def infer_shape_kind(columns: list[str], rows: list[dict[str, Any]]) -> str:
    lowered = " ".join(col.lower() for col in columns)
    if len(columns) >= _WIDE_TABLE_COLUMN_THRESHOLD:
        return "wide_feature_table"
    if ("event" in lowered or "action" in lowered or "status" in lowered) and (
        "time" in lowered or "date" in lowered
    ):
        return "event_log"
    if {"key", "value"}.issubset({col.lower() for col in columns}):
        return "json_kv_table"
    nested = 0
    for row in rows[:50]:
        nested += sum(1 for value in row.values() if isinstance(value, list | dict))
    if rows and nested / max(1, len(rows)) >= 1:
        return "json_kv_table"
    return "flat_table"


def is_number_like(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int | float | Decimal):
        return True
    text = str(value).strip().replace(",", "").replace("楼", "")
    if not text:
        return False
    try:
        Decimal(text)
        return True
    except InvalidOperation:
        return False


def profile_dataset(dataset_id: str, *, owner_id: str, store: StructuredStore | None = None) -> dict[str, Any]:
    store = store or get_structured_store()
    dataset = store.get_dataset(dataset_id, owner_id=owner_id)
    if not dataset:
        raise ValueError("dataset not found")
    table = load_dataset_rows(dataset, owner_id=owner_id, limit=MAX_PROFILE_ROWS, store=store)
    columns = [profile_column(column, [row.get(column) for row in table.rows]) for column in table.columns]
    semantic_result = infer_column_semantics_with_has(dataset, table, columns)
    columns = merge_semantic_column_profiles(columns, semantic_result.get("columns") or {})
    profile = {
        "dataset_id": dataset_id,
        "shape_kind": infer_shape_kind(table.columns, table.rows),
        "row_count_estimate": table.row_count_estimate,
        "sampled_rows": len(table.rows),
        "columns": columns,
        "semantic_inference": {
            "engine": "has_ner",
            "status": semantic_result.get("status", "unknown"),
            "duration_ms": semantic_result.get("duration_ms", 0),
            "matched_columns": len(semantic_result.get("columns") or {}),
        },
    }
    store.save_profile(dataset_id, owner_id=owner_id, profile=profile)
    if not store.get_policy(dataset_id, owner_id=owner_id):
        store.save_policy(dataset_id, owner_id=owner_id, policy=default_policy(profile))
    return profile


def profile_column(column: str, values: list[Any]) -> dict[str, Any]:
    non_empty = [value for value in values if value not in (None, "")]
    total = max(1, len(values))
    null_rate = round(1 - (len(non_empty) / total), 4)
    unique_rate = round(len({normalize_value(value) for value in non_empty}) / max(1, len(non_empty)), 4)
    samples = list(dict.fromkeys(cell_to_json(value) for value in non_empty[:20]))[:8]
    if is_probable_technical_identifier(column, non_empty, unique_rate):
        return {
            "name": column,
            "data_type": infer_runtime_type(non_empty),
            "null_rate": null_rate,
            "unique_rate": unique_rate,
            "sample_values": samples,
            "entity_type": "CUSTOM",
            "risk_level": "low",
            "confidence": 0.88,
            "reasons": ["technical_identifier"],
            "recommended_policy": "keep",
        }
    by_value = classify_by_values(non_empty)
    by_name = classify_by_name(column)
    if is_business_descriptor_column(column) and not blocks_business_descriptor(by_name, by_value):
        return {
            "name": column,
            "data_type": infer_runtime_type(non_empty),
            "null_rate": null_rate,
            "unique_rate": unique_rate,
            "sample_values": samples,
            "entity_type": "CUSTOM",
            "risk_level": "low",
            "confidence": 0.84,
            "reasons": ["business_descriptor"],
            "recommended_policy": "keep",
        }
    chosen = choose_classification(by_name, by_value, unique_rate)
    entity_type, risk_level, confidence, reasons = chosen
    return {
        "name": column,
        "data_type": infer_runtime_type(non_empty),
        "null_rate": null_rate,
        "unique_rate": unique_rate,
        "sample_values": samples,
        "entity_type": entity_type,
        "risk_level": risk_level,
        "confidence": confidence,
        "reasons": reasons,
        "recommended_policy": recommended_action(entity_type, risk_level, unique_rate),
    }


def infer_column_semantics_with_has(
    dataset: dict[str, Any],
    table: LoadedTable,
    columns: list[dict[str, Any]],
) -> dict[str, Any]:
    started = time.perf_counter()
    if not table.rows or not table.columns:
        return {"status": "skipped_empty", "columns": {}, "duration_ms": 0}
    if not has_text_semantic_ready():
        return {"status": "unavailable", "columns": {}, "duration_ms": 0}

    text, samples_by_column = build_structured_ner_text(dataset, table, columns)
    if not samples_by_column:
        return {"status": "skipped_no_candidates", "columns": {}, "duration_ms": 0}
    if not text.strip():
        return {"status": "skipped_empty", "columns": {}, "duration_ms": 0}

    try:
        from app.services.has_client import HaSClient

        timeout = max(1.0, min(float(settings.STRUCTURED_HAS_TIMEOUT), float(settings.HAS_TIMEOUT)))
        result = HaSClient(timeout=timeout, max_retries=0).ner(text, _STRUCTURED_HAS_NER_TYPES)
    except Exception as exc:
        logger.warning("Structured HaS semantic inference failed: %s", exc)
        return {
            "status": "failed",
            "columns": {},
            "duration_ms": max(0, int((time.perf_counter() - started) * 1000)),
        }

    semantic_columns = map_has_entities_to_columns(result, samples_by_column)
    return {
        "status": "used" if semantic_columns else "used_no_matches",
        "columns": semantic_columns,
        "duration_ms": max(0, int((time.perf_counter() - started) * 1000)),
    }


def has_text_semantic_ready() -> bool:
    global _SEMANTIC_READY_CACHE
    now = time.monotonic()
    cached_at, cached_ready = _SEMANTIC_READY_CACHE
    if now - cached_at < _SEMANTIC_READY_TTL_SEC:
        return cached_ready
    try:
        from app.core.config import get_has_health_check_url

        response = httpx.get(get_has_health_check_url(), timeout=2.0, trust_env=False)
        ready = response.status_code < 500
    except Exception:
        ready = False
    _SEMANTIC_READY_CACHE = (now, ready)
    return ready


def build_structured_ner_text(
    dataset: dict[str, Any],
    table: LoadedTable,
    columns: list[dict[str, Any]],
) -> tuple[str, dict[str, list[str]]]:
    profile_by_name = {str(column.get("name")): column for column in columns}
    lines = [
        f"Dataset: {dataset.get('name') or dataset.get('table_name') or 'structured_table'}",
        f"Kind: {dataset.get('source_kind') or 'table'}",
    ]
    samples_by_column: dict[str, list[str]] = {}
    for column in table.columns[:80]:
        profile = profile_by_name.get(column, {})
        raw_samples = list(dict.fromkeys(
            normalize_value(row.get(column))
            for row in table.rows[:80]
            if normalize_value(row.get(column))
        ))[:6]
        samples = [sample[:120] for sample in raw_samples if sample][:6]
        if not samples:
            continue
        if not should_include_column_for_structured_semantics(column, profile, samples):
            continue
        samples_by_column[column] = samples
        lines.append(
            "Column "
            + json.dumps(column, ensure_ascii=False)
            + f" type={profile.get('data_type', 'string')} samples: "
            + " | ".join(samples)
        )
    return "\n".join(lines)[: min(12_000, int(settings.HAS_NER_CONTEXT_TOKENS) * 2)], samples_by_column


def should_include_column_for_structured_semantics(
    column: str,
    profile: dict[str, Any],
    samples: list[str],
) -> bool:
    """Only send genuinely ambiguous table columns to HaS semantic enrichment.

    Deterministic table signals cover most structured PII. Calling the text
    model for every obvious phone/email/name column makes the policy screen feel
    slow and does not improve recall, so HaS is reserved for low-confidence
    natural-language columns where the field name is not enough.
    """
    reasons = {str(reason) for reason in (profile.get("reasons") or [])}
    if reasons.intersection({"technical_identifier", "business_descriptor"}):
        return False
    entity_type = str(profile.get("entity_type") or "CUSTOM")
    confidence = float(profile.get("confidence") or 0)
    if entity_type != "CUSTOM" and confidence >= 0.65:
        return False
    if is_identifier_column_name(column):
        return False
    return any(sample_looks_semantic(sample) for sample in samples)


def is_identifier_column_name(column: str) -> bool:
    text = re.sub(r"[^a-z0-9_\u4e00-\u9fff]+", "_", str(column or "").strip().lower()).strip("_")
    if not text:
        return False
    if text in {"id", "uuid", "guid", "pk", "key", "code", "no", "num", "number", "编号", "序号", "代码"}:
        return True
    return bool(re.search(r"(^|_)(id|uuid|guid|code|no|num|number|编号|序号|代码)$", text))


def sample_looks_semantic(sample: str) -> bool:
    text = normalize_value(sample)
    if not text:
        return False
    if re.search(r"[\u4e00-\u9fff]", text):
        return True
    if re.search(r"[A-Za-z]+[\s路.'-]+[A-Za-z]+", text):
        return True
    if len(text) >= 16 and re.search(r"[A-Za-z]", text) and not re.fullmatch(r"[A-Za-z0-9_\-]+", text):
        return True
    return False


def map_has_entities_to_columns(
    ner_result: dict[str, list[str]],
    samples_by_column: dict[str, list[str]],
) -> dict[str, dict[str, Any]]:
    votes: dict[str, Counter[str]] = {column: Counter() for column in samples_by_column}
    matches: dict[str, dict[str, list[str]]] = {column: {} for column in samples_by_column}
    for raw_type, values in (ner_result or {}).items():
        entity_type = normalize_structured_entity_type(raw_type)
        if entity_type == "CUSTOM" or entity_type not in _STRUCTURED_SEMANTIC_TYPE_RISK:
            continue
        if not isinstance(values, list):
            continue
        for raw_value in values:
            entity_text = compact_text(raw_value)
            if not entity_text:
                continue
            for column, samples in samples_by_column.items():
                if any(entity_matches_sample(entity_text, sample) for sample in samples):
                    votes[column][entity_type] += 1
                    bucket = matches[column].setdefault(entity_type, [])
                    value = normalize_value(raw_value)
                    if value not in bucket:
                        bucket.append(value)
    semantic: dict[str, dict[str, Any]] = {}
    for column, counter in votes.items():
        if not counter:
            continue
        entity_type, count = counter.most_common(1)[0]
        sample_count = max(1, len(samples_by_column.get(column) or []))
        match_ratio = min(1.0, count / sample_count)
        semantic[column] = {
            "entity_type": entity_type,
            "risk_level": _STRUCTURED_SEMANTIC_TYPE_RISK.get(entity_type, "high"),
            "confidence": round(min(0.97, max(0.72, 0.62 + match_ratio * 0.3)), 3),
            "reason": "semantic_model_value",
            "matched_values": matches[column].get(entity_type, [])[:5],
        }
    return semantic


def normalize_structured_entity_type(raw_type: str) -> str:
    value = str(raw_type or "").strip()
    if not value:
        return "CUSTOM"
    mapped = cn_to_id(value)
    canonical = canonical_type_id(mapped)
    if canonical in {"BANK_ACCOUNT"}:
        return "BANK_CARD"
    if canonical in {"PASSWORD"}:
        return "AUTH_SECRET"
    if canonical in _STRUCTURED_SEMANTIC_TYPE_RISK:
        return canonical
    return canonical if canonical in {"PERSON", "PHONE", "EMAIL", "ID_CARD", "ADDRESS"} else "CUSTOM"


def entity_matches_sample(entity_text: str, sample: str) -> bool:
    sample_text = compact_text(sample)
    if not entity_text or not sample_text:
        return False
    if entity_text in sample_text:
        return True
    return len(entity_text) >= 4 and sample_text in entity_text


def merge_semantic_column_profiles(
    columns: list[dict[str, Any]],
    semantic_columns: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    for column in columns:
        out = dict(column)
        semantic = semantic_columns.get(str(column.get("name") or ""))
        if semantic and should_apply_semantic_profile(out, semantic):
            entity_type = str(semantic.get("entity_type") or out.get("entity_type") or "CUSTOM")
            risk_level = str(semantic.get("risk_level") or out.get("risk_level") or "low")
            confidence = float(semantic.get("confidence") or out.get("confidence") or 0)
            reasons = list(out.get("reasons") or [])
            for reason in ["semantic_model", str(semantic.get("reason") or "")]:
                if reason and reason not in reasons:
                    reasons.append(reason)
            out.update(
                {
                    "entity_type": entity_type,
                    "risk_level": risk_level,
                    "confidence": round(max(float(out.get("confidence") or 0), confidence), 3),
                    "reasons": reasons,
                    "recommended_policy": recommended_action(entity_type, risk_level, float(out.get("unique_rate") or 0)),
                }
            )
        merged.append(out)
    return merged


def should_apply_semantic_profile(column: dict[str, Any], semantic: dict[str, Any]) -> bool:
    semantic_type = str(semantic.get("entity_type") or "CUSTOM")
    if semantic_type == "CUSTOM":
        return False
    confidence = float(semantic.get("confidence") or 0)
    current_type = str(column.get("entity_type") or "CUSTOM")
    reasons = {str(reason) for reason in (column.get("reasons") or [])}
    if "technical_identifier" in reasons:
        return False
    if "column_values" in reasons and current_type in _STRUCTURED_DIRECT_VALUE_TYPES:
        return False
    if current_type == semantic_type:
        return confidence >= 0.5
    if current_type == "CUSTOM" or reasons == {"high_cardinality"}:
        return confidence >= _CUSTOM_TYPE_CONFIDENCE_MIN
    return confidence >= _DEFAULT_TYPE_CONFIDENCE_MIN


def classify_by_name(column: str) -> tuple[str, str, float, str] | None:
    text = column.strip().lower()
    for pattern, entity_type, risk_level, confidence in _COLUMN_NAME_HINTS:
        if re.search(pattern, text, re.I):
            return entity_type, risk_level, confidence, "column_name"
    return None


def is_business_descriptor_column(column: str) -> bool:
    text = re.sub(r"[^a-z0-9_\-\s\u4e00-\u9fff]+", "_", str(column or "").strip().lower()).strip("_")
    if not text:
        return False
    return any(pattern.search(text) for pattern in _BUSINESS_DESCRIPTOR_COLUMN_PATTERNS)


def blocks_business_descriptor(
    by_name: tuple[str, str, float, str] | None,
    by_value: tuple[str, str, float, str] | None,
) -> bool:
    for classification in (by_name, by_value):
        if classification and classification[0] in _BUSINESS_DESCRIPTOR_BLOCK_TYPES:
            return True
    return False


def is_probable_technical_identifier(column: str, values: list[Any], unique_rate: float) -> bool:
    normalized = re.sub(r"[^a-z0-9_\u4e00-\u9fff]+", "_", str(column or "").strip().lower()).strip("_")
    if normalized not in {"id", "row_id", "rowid", "pk", "index", "idx", "serial", "\u5e8f\u53f7", "\u884c\u53f7"}:
        return False
    text_values = [normalize_value(value).strip() for value in values if normalize_value(value).strip()]
    if not text_values:
        return False
    if not all(re.fullmatch(r"\d{1,9}", value) for value in text_values):
        return False
    if unique_rate < 0.8:
        return False
    numbers = [int(value) for value in text_values]
    sorted_unique = sorted(set(numbers))
    if len(sorted_unique) == 1:
        return sorted_unique[0] <= max(10_000_000, len(text_values) * 10)
    span = sorted_unique[-1] - sorted_unique[0] + 1
    density = len(sorted_unique) / max(1, span)
    return density >= 0.8 and sorted_unique[-1] <= max(10_000_000, len(text_values) * 10)


def classify_by_values(values: list[Any]) -> tuple[str, str, float, str] | None:
    if not values:
        return None
    sample = [normalize_value(value) for value in values[:200]]
    votes: Counter[tuple[str, str]] = Counter()
    best_confidence = 0.0
    for value in sample:
        text = compact_text(value)
        if not text:
            continue
        for pattern, entity_type, risk_level, confidence in _VALUE_PATTERNS:
            if pattern.match(text):
                votes[(entity_type, risk_level)] += 1
                best_confidence = max(best_confidence, confidence)
                break
    if not votes:
        return None
    (entity_type, risk_level), count = votes.most_common(1)[0]
    ratio = count / max(1, len(sample))
    if ratio < 0.35:
        return None
    return entity_type, risk_level, round(min(0.99, max(best_confidence, ratio)), 3), "column_values"


def choose_classification(
    by_name: tuple[str, str, float, str] | None,
    by_value: tuple[str, str, float, str] | None,
    unique_rate: float,
) -> tuple[str, str, float, list[str]]:
    if by_name and by_value:
        if by_name[0] == by_value[0]:
            return by_name[0], max_risk(by_name[1], by_value[1]), max(by_name[2], by_value[2]), [by_name[3], by_value[3]]
        if by_value[2] >= by_name[2] + 0.12:
            return by_value[0], by_value[1], by_value[2], [by_value[3], "value_overrides_name"]
        return by_name[0], by_name[1], by_name[2], [by_name[3], "name_overrides_value"]
    if by_value:
        return by_value[0], by_value[1], by_value[2], [by_value[3]]
    if by_name:
        return by_name[0], by_name[1], by_name[2], [by_name[3]]
    if unique_rate >= 0.98:
        return "CUSTOM", "low", 0.35, ["high_cardinality"]
    return "CUSTOM", "low", 0.1, []


def max_risk(left: str, right: str) -> str:
    order = {"low": 0, "medium": 1, "high": 2, "critical": 3}
    return left if order.get(left, 0) >= order.get(right, 0) else right


def recommended_action(entity_type: str, risk_level: str, unique_rate: float) -> str:
    del risk_level
    if entity_type in _SECURITY_DEFAULT_SUPPRESS_TYPES:
        return "suppress"
    if entity_type in _PII_DEFAULT_MASK_TYPES:
        return "mask"
    if entity_type in _PII_DEFAULT_HASH_TYPES:
        return "hash"
    if entity_type in _PII_DEFAULT_GENERALIZE_TYPES:
        return "generalize"
    if entity_type in _PII_DEFAULT_TOKENIZE_TYPES:
        return "tokenize" if unique_rate > 0.5 else "mask"
    return "keep"


def default_policy(profile: dict[str, Any]) -> dict[str, Any]:
    return {
        "dataset_id": profile["dataset_id"],
        "columns": [
            {
                "column": col["name"],
                "action": col.get("recommended_policy") or "keep",
                "entity_type": col.get("entity_type") or "CUSTOM",
                "enabled": (col.get("recommended_policy") or "keep") != "keep",
                "params": {},
            }
            for col in profile.get("columns", [])
        ],
    }


def get_or_create_policy(dataset_id: str, *, owner_id: str, store: StructuredStore | None = None) -> dict[str, Any]:
    store = store or get_structured_store()
    policy = store.get_policy(dataset_id, owner_id=owner_id)
    if policy:
        return policy
    profile = store.get_profile(dataset_id, owner_id=owner_id) or profile_dataset(dataset_id, owner_id=owner_id, store=store)
    return store.save_policy(dataset_id, owner_id=owner_id, policy=default_policy(profile))


def save_policy(
    dataset_id: str,
    *,
    owner_id: str,
    columns: list[dict[str, Any]],
    store: StructuredStore | None = None,
) -> dict[str, Any]:
    store = store or get_structured_store()
    if not store.get_dataset(dataset_id, owner_id=owner_id):
        raise ValueError("dataset not found")
    profile = store.get_profile(dataset_id, owner_id=owner_id) or profile_dataset(dataset_id, owner_id=owner_id, store=store)
    validate_policy_columns(profile, columns)
    policy = {
        "dataset_id": dataset_id,
        "columns": columns,
        "reviewed_at": datetime.now(UTC).isoformat(),
    }
    return store.save_policy(dataset_id, owner_id=owner_id, policy=policy)


def validate_policy_columns(profile: dict[str, Any], columns: list[dict[str, Any]]) -> None:
    expected = [str(column.get("name") or "") for column in profile.get("columns", [])]
    expected_set = set(expected)
    seen: set[str] = set()
    duplicates: list[str] = []
    provided: list[str] = []
    for item in columns:
        name = str(item.get("column") or "")
        provided.append(name)
        if name in seen:
            duplicates.append(name)
        seen.add(name)
    provided_set = set(provided)
    unknown = sorted(name for name in provided_set if name not in expected_set)
    missing = [name for name in expected if name not in provided_set]
    if duplicates or unknown or missing:
        parts = []
        if duplicates:
            parts.append("duplicate columns: " + ", ".join(sorted(set(duplicates))[:5]))
        if unknown:
            parts.append("unknown columns: " + ", ".join(unknown[:5]))
        if missing:
            parts.append("missing columns: " + ", ".join(missing[:5]))
        raise ValueError(
            "Column policy does not match the current dataset columns ("
            + "; ".join(parts)
            + "). Regenerate the policy or refresh the dataset before saving."
        )


def preview_dataset(
    dataset_id: str,
    *,
    owner_id: str,
    limit: int = MAX_PREVIEW_ROWS,
    store: StructuredStore | None = None,
) -> dict[str, Any]:
    store = store or get_structured_store()
    dataset = store.get_dataset(dataset_id, owner_id=owner_id)
    if not dataset:
        raise ValueError("dataset not found")
    table = load_dataset_rows(dataset, owner_id=owner_id, limit=min(MAX_PREVIEW_ROWS, max(1, limit)), store=store)
    policy = get_or_create_policy(dataset_id, owner_id=owner_id, store=store)
    policy_columns = list(policy.get("columns") or [])
    return {
        "dataset_id": dataset_id,
        "columns": table.columns,
        "original_rows": table.rows,
        "redacted_rows": [redact_row(row, policy_columns, owner_id=owner_id, dataset_id=dataset_id) for row in table.rows],
        "policy": policy_columns,
    }


def load_dataset_rows(
    dataset: dict[str, Any],
    *,
    owner_id: str,
    limit: int | None,
    store: StructuredStore | None = None,
) -> LoadedTable:
    store = store or get_structured_store()
    source_kind = str(dataset.get("source_kind") or "")
    metadata = dataset.get("metadata") or {}
    if dataset.get("source_id"):
        path = str(metadata.get("path") or "")
        if source_kind == "csv":
            return read_csv(path, limit=limit)
        if source_kind == "jsonl":
            return read_jsonl(path, limit=limit)
        if source_kind == "xlsx":
            return read_xlsx_sheet(path, sheet_name=str(metadata.get("sheet_name") or ""), limit=limit)
        if source_kind == "sqlite":
            return read_sqlite_table(path, table_name=str(dataset.get("table_name") or dataset["name"]), limit=limit)
    if dataset.get("connection_id"):
        connection = store.get_connection(str(dataset["connection_id"]), owner_id=owner_id, include_secret=True)
        if not connection:
            raise ValueError("connection not found")
        credential = decrypt_credential(connection.get("credential") or {})
        return read_connection_table(
            connection,
            credential,
            schema_name=dataset.get("schema_name"),
            table_name=str(dataset.get("table_name") or dataset["name"]),
            limit=limit,
        )
    raise ValueError("dataset source not available")


def read_connection_table(
    connection: dict[str, Any],
    credential: dict[str, Any],
    *,
    schema_name: str | None,
    table_name: str,
    limit: int | None,
) -> LoadedTable:
    engine = str(connection.get("engine") or "")
    if engine == "sqlite":
        path = str(credential.get("sqlite_path") or credential.get("database") or "")
        return read_sqlite_table(path, table_name=table_name, limit=limit)
    sa = sqlalchemy()
    url = build_sqlalchemy_url({**credential, "engine": engine})
    sql_engine = sa.create_engine(url)
    try:
        table_ref = quote_sa_table(sa, sql_engine, schema_name=schema_name, table_name=table_name)
        with sql_engine.connect() as conn:
            empty_result = conn.execute(sa.text(f"SELECT * FROM {table_ref} LIMIT 0"))
            columns = [str(key) for key in empty_result.keys()]
            total = conn.execute(sa.text(f"SELECT COUNT(*) FROM {table_ref}")).scalar_one_or_none()
            sql = f"SELECT * FROM {table_ref}"
            if limit is not None:
                sql += f" LIMIT {int(limit)}"
            rows = [dict(row._mapping) for row in conn.execute(sa.text(sql)).fetchall()]
        return LoadedTable(columns=columns, rows=rows, row_count_estimate=int(total or 0))
    finally:
        sql_engine.dispose()


def sqlalchemy():
    import sqlalchemy as sa

    return sa


def quote_sa_table(sa: Any, engine: Any, *, schema_name: str | None, table_name: str) -> str:
    preparer = engine.dialect.identifier_preparer
    table = preparer.quote(table_name)
    if schema_name:
        return f"{preparer.quote(schema_name)}.{table}"
    del sa
    return table


def _validate_db_host_allowed(host: str) -> None:
    """SSRF guard: when STRUCTURED_DB_HOST_ALLOWLIST is set, reject hosts outside it.

    Allowlist entries are exact hostnames or IP / CIDR networks. ``None`` means
    no restriction (default), preserving the local-tool use case of connecting
    to the user's own databases.
    """
    allowlist = settings.STRUCTURED_DB_HOST_ALLOWLIST
    if allowlist is None:
        return
    try:
        addr = ipaddress.ip_address(host)
    except ValueError:
        addr = None
    for raw_entry in allowlist:
        entry = str(raw_entry).strip()
        if not entry:
            continue
        if host == entry:
            return
        if addr is not None:
            try:
                if addr in ipaddress.ip_network(entry, strict=False):
                    return
            except ValueError:
                continue
    raise ValueError(
        f"database host '{host}' is blocked by STRUCTURED_DB_HOST_ALLOWLIST; "
        "add the exact hostname or an IP/CIDR entry to the allowlist to permit this connection"
    )


def build_sqlalchemy_url(payload: dict[str, Any]) -> str:
    engine = str(payload.get("engine") or "")
    if engine == "mysql":
        driver = "mysql+pymysql"
        port = int(payload.get("port") or 3306)
    elif engine == "postgres":
        driver = "postgresql+psycopg"
        port = int(payload.get("port") or 5432)
    elif engine == "sqlite":
        path = str(payload.get("sqlite_path") or payload.get("database") or "")
        return f"sqlite:///{Path(path).as_posix()}"
    else:
        raise ValueError(f"unsupported database engine: {engine}")
    username = str(payload.get("username") or "")
    password = str(payload.get("password") or "")
    host = str(payload.get("host") or "localhost")
    _validate_db_host_allowed(host)
    database = str(payload.get("database") or "")
    from urllib.parse import quote_plus

    return f"{driver}://{quote_plus(username)}:{quote_plus(password)}@{host}:{port}/{quote_plus(database)}"


def test_connection(payload: dict[str, Any]) -> dict[str, Any]:
    datasets = discover_connection_datasets_from_payload(payload, limit=50)
    return {
        "ok": True,
        "message": "connection ok",
        "engine": payload.get("engine"),
        "dataset_count": len(datasets),
    }


def connection_display_metadata(payload: dict[str, Any], *, dataset_count: int) -> dict[str, Any]:
    """Return non-secret connection details that help users identify saved targets."""
    engine = str(payload.get("engine") or "")
    metadata: dict[str, Any] = {
        "dataset_count": int(dataset_count),
    }
    if engine == "sqlite":
        sqlite_path = str(payload.get("sqlite_path") or payload.get("database") or "").strip()
        if sqlite_path:
            metadata["sqlite_path"] = sqlite_path
            metadata["target"] = sqlite_path
        return metadata

    host = str(payload.get("host") or "").strip()
    port = payload.get("port")
    database = str(payload.get("database") or "").strip()
    username = str(payload.get("username") or "").strip()
    if host:
        metadata["host"] = host
    if port:
        metadata["port"] = int(port)
    if database:
        metadata["database"] = database
    if username:
        metadata["username"] = username
    endpoint = host
    if port:
        endpoint = f"{endpoint}:{int(port)}" if endpoint else str(int(port))
    if database:
        endpoint = f"{endpoint}/{database}" if endpoint else database
    if endpoint:
        metadata["target"] = endpoint
    return metadata


def create_connection(
    *,
    owner_id: str,
    payload: dict[str, Any],
    store: StructuredStore | None = None,
) -> dict[str, Any]:
    store = store or get_structured_store()
    test = test_connection(payload)
    connection = store.create_connection(
        owner_id=owner_id,
        engine=str(payload.get("engine")),
        display_name=str(payload.get("display_name") or payload.get("database") or payload.get("sqlite_path") or "Database"),
        encrypted_credential=encrypt_credential(payload),
        last_test_status="ok" if test["ok"] else "failed",
        metadata=connection_display_metadata(payload, dataset_count=int(test["dataset_count"])),
    )
    return connection


def discover_connection_datasets(
    connection_id: str,
    *,
    owner_id: str,
    store: StructuredStore | None = None,
) -> list[dict[str, Any]]:
    store = store or get_structured_store()
    connection = store.get_connection(connection_id, owner_id=owner_id, include_secret=True)
    if not connection:
        raise ValueError("connection not found")
    credential = decrypt_credential(connection.get("credential") or {})
    datasets = discover_connection_datasets_from_payload({**credential, "engine": connection["engine"]})
    return annotate_discovered_connection_datasets(datasets, connection_id=connection_id)


def register_connection_datasets(
    connection_id: str,
    *,
    owner_id: str,
    selections: list[dict[str, Any]],
    store: StructuredStore | None = None,
) -> list[dict[str, Any]]:
    store = store or get_structured_store()
    connection = store.get_connection(connection_id, owner_id=owner_id, include_secret=True)
    if not connection:
        raise ValueError("connection not found")
    discovered = discover_connection_datasets(connection_id, owner_id=owner_id, store=store)
    by_key = {
        (item.get("schema_name"), item.get("table_name") or item.get("name")): item for item in discovered
    }
    selected: list[dict[str, Any]] = []
    for raw in selections:
        key = (raw.get("schema_name"), raw.get("table_name") or raw.get("name"))
        item = by_key.get(key)
        if not item:
            continue
        selected.append(
            store.upsert_dataset(
                owner_id=owner_id,
                connection_id=connection_id,
                source_id=None,
                name=item["name"],
                dataset_type=item["dataset_type"],
                source_kind=str(connection["engine"]),
                shape_kind=item.get("shape_kind") or "flat_table",
                schema_name=item.get("schema_name"),
                table_name=item.get("table_name"),
                row_count_estimate=item.get("row_count_estimate"),
                column_count=int(item.get("column_count") or 0),
                schema=item.get("schema") or [],
                metadata=item.get("metadata") or {},
            )
        )
    return selected


def discover_connection_datasets_from_payload(payload: dict[str, Any], *, limit: int = _DEFAULT_DATASET_DISCOVERY_LIMIT) -> list[dict[str, Any]]:
    engine = str(payload.get("engine") or "")
    if engine == "sqlite":
        path = str(payload.get("sqlite_path") or payload.get("database") or "")
        if not path or not os.path.exists(path):
            raise ValueError("sqlite database path not found")
        return discover_sqlite_datasets(path, source_id=None, source_kind="sqlite")[:limit]
    sa = sqlalchemy()
    sql_engine = sa.create_engine(build_sqlalchemy_url(payload))
    datasets: list[dict[str, Any]] = []
    try:
        inspector = sa.inspect(sql_engine)
        for schema_name in inspector.get_schema_names():
            if schema_name in {"information_schema", "pg_catalog", "mysql", "performance_schema", "sys"}:
                continue
            table_names = inspector.get_table_names(schema=schema_name)
            view_names = inspector.get_view_names(schema=schema_name)
            for table_name, dataset_type in [(name, "db_table") for name in table_names] + [
                (name, "db_view") for name in view_names
            ]:
                columns = inspector.get_columns(table_name, schema=schema_name)
                schema = [{"name": col["name"], "data_type": str(col.get("type") or "string")} for col in columns]
                datasets.append(
                    {
                        "source_id": None,
                        "connection_id": None,
                        "name": f"{schema_name}.{table_name}" if schema_name else table_name,
                        "dataset_type": dataset_type,
                        "source_kind": engine,
                        "shape_kind": "flat_table",
                        "schema_name": schema_name,
                        "table_name": table_name,
                        "row_count_estimate": None,
                        "column_count": len(schema),
                        "schema": schema,
                        "metadata": {},
                    }
                )
                if len(datasets) >= limit:
                    return datasets
    finally:
        sql_engine.dispose()
    return datasets


def redact_row(row: dict[str, Any], policy_columns: list[dict[str, Any]], *, owner_id: str, dataset_id: str) -> dict[str, Any]:
    by_column = {str(item.get("column")): item for item in policy_columns if item.get("enabled", True)}
    out = dict(row)
    for column, policy in by_column.items():
        if column not in out:
            continue
        out[column] = redact_value(
            out[column],
            action=str(policy.get("action") or "keep"),
            entity_type=str(policy.get("entity_type") or "CUSTOM"),
            salt=f"{owner_id}:{dataset_id}:{column}",
            params=policy.get("params") if isinstance(policy.get("params"), dict) else {},
        )
    return out


def redact_value(value: Any, *, action: str, entity_type: str, salt: str, params: dict[str, Any]) -> Any:
    if value is None or action == "keep":
        return value
    text = normalize_value(value)
    if text == "":
        return value
    if action == "suppress":
        return None
    if action == "hash":
        return stable_digest(text, salt=salt)
    if action == "tokenize":
        return f"{entity_type}_{stable_digest(text, salt=salt)[:12]}"
    if action == "generalize":
        return generalize_value(text, entity_type=entity_type)
    if action == "bucket":
        return bucket_value(text)
    if action == "custom":
        return params.get("replacement", "***")
    return mask_value(text, entity_type=entity_type)


def stable_digest(text: str, *, salt: str) -> str:
    key = (settings.JWT_SECRET_KEY or "structured-redaction").encode("utf-8")
    msg = f"{salt}:{text}".encode()
    return hmac.new(key, msg, hashlib.sha256).hexdigest()[:24]


def mask_value(text: str, *, entity_type: str) -> str:
    compact = compact_text(text)
    if entity_type == "EMAIL" and "@" in text:
        name, domain = text.split("@", 1)
        return f"{mask_keep_edges(name, 1, 0)}@{domain}"
    if entity_type in {"PHONE", "BANK_CARD", "ID_CARD"}:
        return mask_keep_edges(compact, 3, 4)
    if len(text) <= 2:
        return "*" * len(text)
    return mask_keep_edges(text, 1, 1)


def mask_keep_edges(text: str, left: int, right: int) -> str:
    if len(text) <= left + right:
        return "*" * len(text)
    return f"{text[:left]}{'*' * (len(text) - left - right)}{text[-right:] if right else ''}"


def generalize_value(text: str, *, entity_type: str) -> str:
    if entity_type == "DATE":
        match = re.match(r"^(\d{4})[-/骞碷(\d{1,2})", text)
        if match:
            return f"{match.group(1)}-{int(match.group(2)):02d}"
    if entity_type == "ADDRESS":
        return text[:6] + "***" if len(text) > 6 else "***"
    return "***"


def bucket_value(text: str) -> str:
    try:
        amount = Decimal(text.replace(",", "").replace("楼", ""))
    except InvalidOperation:
        return "***"
    if amount < 1_000:
        return "<1k"
    if amount < 10_000:
        return "1k-10k"
    if amount < 100_000:
        return "10k-100k"
    if amount < 1_000_000:
        return "100k-1m"
    return ">=1m"


class _ExportRowLimitExceeded(Exception):
    pass


def copy_stream_with_limit(src, dst_path: str, max_bytes: int, chunk_size: int = 1024 * 1024) -> int:
    """分块拷贝上传流并强制大小上限；超限删除半成品并报错（对齐 /files/upload）。"""
    written = 0
    try:
        with open(dst_path, "wb") as dst:
            while True:
                chunk = src.read(chunk_size)
                if not chunk:
                    break
                written += len(chunk)
                if written > max_bytes:
                    raise ValueError(f"文件超过上传上限 {max_bytes // (1024 * 1024)}MB")
                dst.write(chunk)
    except Exception:
        if os.path.exists(dst_path):
            os.remove(dst_path)
        raise
    return written


def iter_dataset_rows(
    dataset: dict[str, Any],
    *,
    owner_id: str,
    store: StructuredStore | None = None,
):
    """流式读数据集：返回 (columns, 行迭代器)，内存 O(1)。

    preview/profile 仍走 load_dataset_rows（有限行）；导出用本函数，
    十万行不再全量进内存。列集合与既有 read_* 的口径一致。
    """
    store = store or get_structured_store()
    source_kind = str(dataset.get("source_kind") or "")
    metadata = dataset.get("metadata") or {}
    if dataset.get("source_id"):
        path = str(metadata.get("path") or "")
        if source_kind == "csv":
            return _iter_csv_rows(path)
        if source_kind == "jsonl":
            return _iter_jsonl_rows(path)
        if source_kind == "xlsx":
            return _iter_xlsx_rows(path, sheet_name=str(metadata.get("sheet_name") or ""))
        if source_kind == "sqlite":
            return _iter_sqlite_rows(path, table_name=str(dataset.get("table_name") or dataset["name"]))
    if dataset.get("connection_id"):
        connection = store.get_connection(str(dataset["connection_id"]), owner_id=owner_id, include_secret=True)
        if not connection:
            raise ValueError("connection not found")
        credential = decrypt_credential(connection.get("credential") or {})
        engine = str(connection.get("engine") or "")
        table_name = str(dataset.get("table_name") or dataset["name"])
        if engine == "sqlite":
            path = str(credential.get("sqlite_path") or credential.get("database") or "")
            return _iter_sqlite_rows(path, table_name=table_name)
        return _iter_connection_rows(
            connection, credential,
            schema_name=dataset.get("schema_name"), table_name=table_name,
        )
    raise ValueError("dataset source not available")


def _iter_csv_rows(path: str):
    encoding = detect_csv_encoding(path)
    with open(path, encoding=encoding, newline="") as fh:
        sample = fh.read(8192)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        fh.seek(0)
        columns = normalize_columns(csv.DictReader(fh, dialect=dialect).fieldnames or [])

    def rows():
        with open(path, encoding=encoding, newline="") as handle:
            reader = csv.DictReader(handle, dialect=dialect)
            for raw in reader:
                yield {columns[i]: value for i, value in enumerate(raw.values()) if i < len(columns)}

    return columns, rows()


def _iter_jsonl_rows(path: str, probe_limit: int = MAX_PROFILE_ROWS):
    columns: list[str] = []
    seen: set[str] = set()
    with open(path, encoding="utf-8-sig") as fh:
        for index, line in enumerate(fh):
            if index >= probe_limit:
                break
            if not line.strip():
                continue
            obj = json.loads(line)
            if not isinstance(obj, dict):
                obj = {"value": obj}
            for key in obj:
                if key not in seen:
                    seen.add(key)
                    columns.append(str(key))
    normalized = normalize_columns(columns)

    def rows():
        with open(path, encoding="utf-8-sig") as handle:
            for line in handle:
                if not line.strip():
                    continue
                obj = json.loads(line)
                if not isinstance(obj, dict):
                    obj = {"value": obj}
                yield {str(key): value for key, value in obj.items()}

    return normalized, rows()


def _iter_xlsx_rows(path: str, *, sheet_name: str):
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        header = next(ws.iter_rows(values_only=True), None)
        columns = normalize_columns([stringify_cell(value) for value in (header or [])])
        if not columns:
            columns = [f"column_{i + 1}" for i in range(ws.max_column or 0)]
    finally:
        wb.close()

    def rows():
        book = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = book[sheet_name]
            iterator = sheet.iter_rows(values_only=True)
            next(iterator, None)  # skip header
            for values in iterator:
                if not values or all(value is None for value in values):
                    continue
                yield {columns[i]: cell_to_json(value) for i, value in enumerate(values[: len(columns)])}
        finally:
            book.close()

    return columns, rows()


def _iter_sqlite_rows(path: str, *, table_name: str):
    with sqlite3.connect(path) as conn:
        columns = [str(row[1]) for row in conn.execute(f"PRAGMA table_info({quote_sqlite_ident(table_name)})")]

    def rows():
        with sqlite3.connect(path) as conn:
            conn.row_factory = sqlite3.Row
            for row in conn.execute(f"SELECT * FROM {quote_sqlite_ident(table_name)}"):
                yield dict(row)

    return columns, rows()


def _iter_connection_rows(connection, credential, *, schema_name, table_name):
    sa = sqlalchemy()
    engine = str(connection.get("engine") or "")
    url = build_sqlalchemy_url({**credential, "engine": engine})
    sql_engine = sa.create_engine(url)
    try:
        table_ref = quote_sa_table(sa, sql_engine, schema_name=schema_name, table_name=table_name)
        with sql_engine.connect() as conn:
            columns = [str(key) for key in conn.execute(sa.text(f"SELECT * FROM {table_ref} LIMIT 0")).keys()]
    except Exception:
        sql_engine.dispose()
        raise

    def rows():
        try:
            with sql_engine.connect() as conn:
                result = conn.execution_options(stream_results=True, yield_per=1000).execute(
                    sa.text(f"SELECT * FROM {table_ref}")
                )
                for row in result:
                    yield dict(row._mapping)
        finally:
            sql_engine.dispose()

    return columns, rows()


def _write_export_parts(
    fmt: str,
    out_dir: str,
    base_filename: str,
    table_name: str,
    columns: list[str],
    rows_iter,
    rows_per_part: int | None,
) -> list[dict[str, Any]]:
    """流式写导出文件；xlsx 按 rows_per_part 分卷。失败时清掉已写文件再抛。"""
    stem, ext = os.path.splitext(base_filename)
    written_paths: list[str] = []
    parts: list[dict[str, Any]] = []

    def _part_path(index: int) -> tuple[str, str]:
        name = f"{stem}-part{index:03d}{ext}"
        return name, os.path.join(out_dir, name)

    try:
        if fmt == "csv":
            path = os.path.join(out_dir, base_filename)
            written_paths.append(path)
            count = 0
            with open(path, "w", encoding="utf-8-sig", newline="") as fh:
                writer = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
                writer.writeheader()
                for row in rows_iter:
                    writer.writerow(row)
                    count += 1
            parts.append({"filename": base_filename, "path": path, "rows": count})
        elif fmt == "xlsx":
            from openpyxl import Workbook

            cap = max(1, int(rows_per_part or 50_000))
            wb = ws = None
            part_rows = 0

            def _open(index: int):
                nonlocal wb, ws, part_rows
                name, path = _part_path(index)
                written_paths.append(path)
                parts.append({"filename": name, "path": path, "rows": 0})
                wb = Workbook(write_only=True)
                ws = wb.create_sheet("redacted")
                ws.append(columns)
                part_rows = 0

            def _save():
                nonlocal wb
                if wb is not None:
                    wb.save(parts[-1]["path"])
                    parts[-1]["rows"] = part_rows
                    wb = None

            _open(1)
            for row in rows_iter:
                if part_rows >= cap:
                    _save()
                    _open(len(parts) + 1)
                ws.append([row.get(col) for col in columns])
                part_rows += 1
            _save()
            if len(parts) == 1:
                # 单卷去 -part 后缀，保持既有交付命名
                plain_path = os.path.join(out_dir, base_filename)
                os.replace(parts[0]["path"], plain_path)
                parts[0]["filename"] = base_filename
                parts[0]["path"] = plain_path
        elif fmt == "sqlite":
            path = os.path.join(out_dir, base_filename)
            written_paths.append(path)
            if os.path.exists(path):
                os.remove(path)
            count = 0
            with sqlite3.connect(path) as conn:
                cols = ", ".join(f"{quote_sqlite_ident(col)} TEXT" for col in columns)
                conn.execute(f"CREATE TABLE {quote_sqlite_ident(table_name)} ({cols})")
                placeholders = ", ".join("?" for _ in columns)
                col_names = ", ".join(quote_sqlite_ident(col) for col in columns)
                batch: list[list[str]] = []
                for row in rows_iter:
                    batch.append([normalize_value(row.get(col)) for col in columns])
                    count += 1
                    if len(batch) >= 1000:
                        conn.executemany(
                            f"INSERT INTO {quote_sqlite_ident(table_name)} ({col_names}) VALUES ({placeholders})",
                            batch,
                        )
                        batch = []
                if batch:
                    conn.executemany(
                        f"INSERT INTO {quote_sqlite_ident(table_name)} ({col_names}) VALUES ({placeholders})",
                        batch,
                    )
                conn.commit()
            parts.append({"filename": base_filename, "path": path, "rows": count})
        elif fmt == "sql":
            path = os.path.join(out_dir, base_filename)
            written_paths.append(path)
            count = 0
            with open(path, "w", encoding="utf-8") as fh:
                cols = ", ".join(quote_sqlite_ident(col) for col in columns)
                fh.write(f"-- Redacted export generated at {utc_iso()}\n")
                for row in rows_iter:
                    values = ", ".join(sql_literal(row.get(col)) for col in columns)
                    fh.write(f"INSERT INTO {quote_sqlite_ident(table_name)} ({cols}) VALUES ({values});\n")
                    count += 1
            parts.append({"filename": base_filename, "path": path, "rows": count})
        else:
            raise ValueError(f"unsupported export format: {fmt}")
    except Exception:
        for path in written_paths:
            try:
                if os.path.exists(path):
                    os.remove(path)
            except OSError:
                pass
        raise
    return parts


def export_dataset(
    dataset_id: str,
    *,
    owner_id: str,
    job_id: str,
    export_format: str,
    store: StructuredStore | None = None,
) -> dict[str, Any]:
    store = store or get_structured_store()
    dataset = store.get_dataset(dataset_id, owner_id=owner_id)
    if not dataset:
        raise ValueError("dataset not found")
    policy = get_or_create_policy(dataset_id, owner_id=owner_id, store=store)
    policy_columns = policy.get("columns") or []
    fmt = "csv" if export_format == "zip" else export_format
    if fmt not in {"csv", "xlsx", "sqlite", "sql"}:
        raise ValueError(f"unsupported export format: {export_format}")
    out_dir = os.path.join(settings.OUTPUT_DIR, "structured", safe_filename(owner_id), job_id)
    os.makedirs(out_dir, exist_ok=True)
    base = export_base_filename(dataset.get("name") or dataset_id)

    columns, source_rows = iter_dataset_rows(dataset, owner_id=owner_id, store=store)
    max_rows = int(getattr(settings, "STRUCTURED_MAX_EXPORT_ROWS", 0) or MAX_EXPORT_ROWS)
    counter = {"rows": 0}

    def redacted_rows():
        for row in source_rows:
            counter["rows"] += 1
            if counter["rows"] > max_rows:
                raise _ExportRowLimitExceeded()
            yield redact_row(row, policy_columns, owner_id=owner_id, dataset_id=dataset_id)

    rows_per_part = int(getattr(settings, "EXPORT_TABLE_ROWS_PER_FILE", 50_000)) if fmt == "xlsx" else None

    with _STRUCTURED_EXPORT_LOCK:
        filename = unique_export_filename(
            f"{base}.{fmt}",
            dataset_id=dataset_id,
            owner_id=owner_id,
            job_id=job_id,
            out_dir=out_dir,
            store=store,
        )
        export_base = export_base_filename(filename)
        try:
            parts = _write_export_parts(
                fmt, out_dir, filename, export_base, columns, redacted_rows(), rows_per_part
            )
        except _ExportRowLimitExceeded:
            # 绝不静默截断：超限即失败，partial 已被 _write_export_parts 清理
            raise ValueError(
                f"数据集超过导出上限 {max_rows} 行（实际 ≥{counter['rows']}）；"
                "请分表导出或调高 STRUCTURED_MAX_EXPORT_ROWS"
            )

    total_rows = counter["rows"]
    action_counts = Counter(str(col.get("action") or "keep") for col in policy_columns if col.get("enabled", True))
    redacted_columns = sum(1 for col in policy_columns if col.get("enabled", True) and col.get("action") != "keep")
    base_summary = {
        "dataset_id": dataset_id,
        "dataset_name": dataset.get("name"),
        "row_count": total_rows,
        "column_count": len(columns),
        "redacted_column_count": redacted_columns,
        "action_counts": dict(action_counts),
        "shape_kind": dataset.get("shape_kind"),
        "part_count": len(parts),
        "total_rows": total_rows,
    }
    record: dict[str, Any] | None = None
    for index, part in enumerate(parts):
        record = store.add_export(
            owner_id=owner_id,
            job_id=job_id,
            dataset_id=dataset_id,
            export_format=fmt,
            file_path=part["path"],
            filename=part["filename"],
            summary={
                **base_summary,
                "export_filename": part["filename"],
                "part_index": index + 1,
                "part_rows": part["rows"],
            },
        )
    assert record is not None
    return record


def write_csv(path: str, columns: list[str], rows: list[dict[str, Any]]) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: str, columns: list[str], rows: list[dict[str, Any]]) -> None:
    from openpyxl import Workbook

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("redacted")
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col) for col in columns])
    wb.save(path)


def write_sqlite(path: str, *, table_name: str, columns: list[str], rows: list[dict[str, Any]]) -> None:
    if os.path.exists(path):
        os.remove(path)
    with sqlite3.connect(path) as conn:
        cols = ", ".join(f"{quote_sqlite_ident(col)} TEXT" for col in columns)
        conn.execute(f"CREATE TABLE {quote_sqlite_ident(table_name)} ({cols})")
        placeholders = ", ".join("?" for _ in columns)
        col_names = ", ".join(quote_sqlite_ident(col) for col in columns)
        conn.executemany(
            f"INSERT INTO {quote_sqlite_ident(table_name)} ({col_names}) VALUES ({placeholders})",
            [[normalize_value(row.get(col)) for col in columns] for row in rows],
        )
        conn.commit()


def write_sql(path: str, *, table_name: str, columns: list[str], rows: list[dict[str, Any]]) -> None:
    with open(path, "w", encoding="utf-8") as fh:
        cols = ", ".join(quote_sqlite_ident(col) for col in columns)
        fh.write(f"-- Redacted export generated at {utc_iso()}\n")
        for row in rows:
            values = ", ".join(sql_literal(row.get(col)) for col in columns)
            fh.write(f"INSERT INTO {quote_sqlite_ident(table_name)} ({cols}) VALUES ({values});\n")


def sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    return "'" + normalize_value(value).replace("'", "''") + "'"


async def run_structured_job_item(
    *,
    job_id: str,
    item_id: str,
    dataset_id: str,
    owner_id: str,
    export_format: str,
    store: Any,
) -> dict[str, Any]:
    started = time.perf_counter()
    store.update_item_progress(
        item_id,
        stage="structured_profile",
        current=1,
        total=3,
        message="Profiling table columns",
    )
    profile = profile_dataset(dataset_id, owner_id=owner_id)
    store.update_item_progress(
        item_id,
        stage="structured_redaction",
        current=2,
        total=3,
        message="Applying column policy",
    )
    export = export_dataset(dataset_id, owner_id=owner_id, job_id=job_id, export_format=export_format)
    store.update_item_progress(
        item_id,
        stage="structured_export",
        current=3,
        total=3,
        message="Structured export ready",
    )
    return {
        "profile": {
            "sampled_rows": profile.get("sampled_rows"),
            "column_count": len(profile.get("columns") or []),
        },
        "export": export,
        "duration_ms": max(0, int((time.perf_counter() - started) * 1000)),
    }


def build_job_export_zip(*, owner_id: str, job_id: str, store: StructuredStore | None = None) -> str:
    store = store or get_structured_store()
    exports = store.list_exports(owner_id=owner_id, job_id=job_id)
    if not exports:
        raise ValueError("no structured exports ready")
    out_dir = os.path.join(settings.OUTPUT_DIR, "structured", safe_filename(owner_id), job_id)
    os.makedirs(out_dir, exist_ok=True)
    zip_path = os.path.join(out_dir, "structured-redacted.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        manifest = {
            "job_id": job_id,
            "generated_at": utc_iso(),
            "exports": [
                {
                    "dataset_id": export.get("dataset_id"),
                    "filename": export.get("filename"),
                    "summary": export.get("summary") or {},
                }
                for export in exports
            ],
        }
        for export in exports:
            file_path = str(export.get("file_path") or "")
            if os.path.isfile(file_path):
                zf.write(file_path, arcname=str(export.get("filename") or os.path.basename(file_path)))
        zf.writestr("quality-report.json", json.dumps(manifest, ensure_ascii=False, indent=2))
    return zip_path


def structured_item_meta(dataset_id: str, *, owner_id: str, store: StructuredStore | None = None) -> dict[str, Any] | None:
    store = store or get_structured_store()
    dataset = store.get_dataset(dataset_id, owner_id=owner_id)
    if not dataset:
        return None
    policy = store.get_policy(dataset_id, owner_id=owner_id) or {}
    redacted_columns = sum(1 for col in policy.get("columns") or [] if col.get("enabled", True) and col.get("action") != "keep")
    return {
        "filename": dataset.get("name"),
        "file_type": f"structured:{dataset.get('source_kind')}",
        "has_output": False,
        "entity_count": redacted_columns,
        "metadata_warning": None,
    }


def normalize_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def compact_text(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_value(value))


def base64_preview(value: bytes) -> str:
    return base64.b64encode(value[:24]).decode("ascii")
