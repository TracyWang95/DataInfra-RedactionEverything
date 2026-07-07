"""库表导出流式化 + 分卷 + 去静默截断（P0-表格）。

背景（2026-07 "10 万条导出表格没法打开"）：旧 export_dataset 全量进内存、
xlsx 单 sheet 十万行 Excel 卡死、超 MAX_EXPORT_ROWS(25 万) 静默截断丢数据。
钉住新行为：
  - 超上限报错且不留 partial、不产生 export 记录（绝不静默丢数据）
  - xlsx 按 EXPORT_TABLE_ROWS_PER_FILE 分卷，每卷一条 export 记录
  - 小表单文件名不带 -part 后缀（兼容既有交付命名）
  - csv 流式结果与旧全量路径等价
"""

import csv
import io
from pathlib import Path

import pytest

import app.services.structured_service as structured_service
from app.services.structured_store import StructuredStore


def _write_rows_csv(path: Path, count: int) -> None:
    with io.open(path, "w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["customer_name", "mobile_phone", "note"])
        for i in range(count):
            writer.writerow([f"客户{i}", f"138{i:08d}", f"备注{i}"])


@pytest.fixture()
def dataset_env(tmp_path, monkeypatch):
    monkeypatch.setattr(structured_service.settings, "OUTPUT_DIR", str(tmp_path / "output"))
    monkeypatch.setattr(structured_service.settings, "DATA_DIR", str(tmp_path / "data"))
    monkeypatch.setattr(structured_service.settings, "JWT_SECRET_KEY", "test-secret")
    store = StructuredStore(str(tmp_path / "structured.sqlite3"))

    def register(count: int, name: str = "big_table.csv") -> str:
        csv_path = tmp_path / name
        _write_rows_csv(csv_path, count)
        result = structured_service.register_file_source(
            owner_id="tenant_a",
            filename=name,
            file_path=str(csv_path),
            kind="csv",
            store=store,
        )
        return result["datasets"][0]["id"]

    return store, register


def _export(store, dataset_id, fmt):
    return structured_service.export_dataset(
        dataset_id,
        owner_id="tenant_a",
        job_id="job_x",
        export_format=fmt,
        store=store,
    )


def test_over_limit_raises_and_leaves_no_partial(dataset_env, monkeypatch, tmp_path):
    store, register = dataset_env
    monkeypatch.setattr(structured_service.settings, "STRUCTURED_MAX_EXPORT_ROWS", 100, raising=False)
    dataset_id = register(250)

    with pytest.raises(ValueError, match="导出上限"):
        _export(store, dataset_id, "csv")

    out_root = tmp_path / "output" / "structured"
    leftovers = list(out_root.rglob("*.csv")) if out_root.exists() else []
    assert leftovers == [], f"partial files left behind: {leftovers}"
    assert store.list_exports(owner_id="tenant_a", job_id="job_x") == []


def test_xlsx_splits_into_parts_with_export_record_each(dataset_env, monkeypatch):
    store, register = dataset_env
    monkeypatch.setattr(structured_service.settings, "EXPORT_TABLE_ROWS_PER_FILE", 100, raising=False)
    dataset_id = register(250)

    last = _export(store, dataset_id, "xlsx")
    exports = store.list_exports(owner_id="tenant_a", job_id="job_x")
    assert len(exports) == 3
    assert last["summary"]["part_count"] == 3
    assert last["summary"]["total_rows"] == 250

    from openpyxl import load_workbook

    part_rows = []
    for record in exports:
        wb = load_workbook(record["file_path"], read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        assert rows[0] == ("customer_name", "mobile_phone", "note")
        part_rows.append(len(rows) - 1)
    assert sorted(part_rows) == [50, 100, 100]
    assert any("-part" in record["filename"] for record in exports)


def test_small_table_keeps_plain_filename(dataset_env):
    store, register = dataset_env
    dataset_id = register(30, name="small.csv")
    export = _export(store, dataset_id, "xlsx")
    assert "-part" not in export["filename"]
    assert export["summary"]["part_count"] == 1
    assert export["summary"]["total_rows"] == 30


def test_csv_streaming_row_count_and_header(dataset_env):
    store, register = dataset_env
    dataset_id = register(250, name="plain.csv")
    export = _export(store, dataset_id, "csv")
    with io.open(export["file_path"], encoding="utf-8-sig") as fh:
        got = list(csv.DictReader(fh))
    assert len(got) == 250
    assert set(got[0].keys()) == {"customer_name", "mobile_phone", "note"}
    assert export["summary"]["row_count"] == 250


def test_copy_stream_with_limit_rejects_oversize(tmp_path):
    src = io.BytesIO(b"x" * 2048)
    dst = tmp_path / "upload.bin"
    with pytest.raises(ValueError, match="上限"):
        structured_service.copy_stream_with_limit(src, str(dst), max_bytes=1024)
    assert not dst.exists()

    src2 = io.BytesIO(b"y" * 512)
    written = structured_service.copy_stream_with_limit(src2, str(dst), max_bytes=1024)
    assert written == 512 and dst.read_bytes() == b"y" * 512
