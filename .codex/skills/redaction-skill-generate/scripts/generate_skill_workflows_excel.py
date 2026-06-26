#!/usr/bin/env python3
"""Generate docs/redaction-skill-workflows.xlsx from workflow stage definitions."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[4]
OUT = ROOT / "docs" / "redaction-skill-workflows.xlsx"

COLUMNS = [
    "业务大类",
    "业务流程",
    "流程总调Skill",
    "阶段序号",
    "阶段名称",
    "阶段类型",
    "Skill ID",
    "Skill调用",
    "输入产出摘要",
    "Cursor提问示例",
]


ROWS: list[tuple] = []

def add(business, flow, orchestrator, step_no, step_name, step_type, skill_id, invoke, io, prompt):
    ROWS.append((business, flow, orchestrator, step_no, step_name, step_type, skill_id, invoke, io, prompt))

# Image flow - module stages
O = "$redaction-anonymize-image-flow"
F = "非结构化单文件 · 图片/扫描页"
add("非结构化", F, O, "—", "流程总调", "总调", "redaction-anonymize-image-flow", O,
    "API 端到端图片脱敏", "请使用 $redaction-anonymize-image-flow 说明 unittest/333.jpg 的模块阶段链（无 UI、无行业预设）。")
add("非结构化", F, O, "0", "服务检查", "子步骤", "redaction-model-service-check", "$redaction-model-service-check",
    "OCR/NER/视觉是否在线", "使用 $redaction-model-service-check 检查模型服务状态。")
add("非结构化", F, O, "1", "OCR模块", "子步骤", "redaction-ocr-module", "$redaction-ocr-module",
    "图片→规范化 OCR blocks", "使用 $redaction-ocr-module 对 unittest/333.jpg 做 OCR，直接返回规范化 blocks。")
add("非结构化", F, O, "2", "文本实体模块", "子步骤", "redaction-text-entity-module", "$redaction-text-entity-module",
    "文本→实体（自选 type）", "使用 $redaction-text-entity-module，entity_types 传 PERSON,INSTITUTION_NAME,ADDRESS,DATE，识别判决书实体。")
add("非结构化", F, O, "3", "实体落框", "子步骤", "redaction-entity-box-map", "$redaction-entity-box-map",
    "实体映射到字形框", "使用 $redaction-entity-box-map 把实体落到 OCR 字形框。")
add("非结构化", F, O, "4", "视觉检测模块", "子步骤", "redaction-visual-detect-module", "$redaction-visual-detect-module",
    "视觉区域+印章+码区", "使用 $redaction-visual-detect-module 检测签字和印章区域。")
add("非结构化", F, O, "5", "区域融合", "子步骤", "redaction-region-deduplicate", "$redaction-region-deduplicate",
    "合并去重候选框", "使用 $redaction-region-deduplicate 合并文本框与视觉框。")
add("非结构化", F, O, "6", "脱敏计划", "子步骤", "redaction-mask-plan-build", "$redaction-mask-plan-build",
    "生成 mask plan", "使用 $redaction-mask-plan-build 生成脱敏计划。")
add("非结构化", F, O, "7", "预览/渲染", "子步骤", "redaction-mask-image-render", "$redaction-mask-image-render",
    "输出脱敏图片", "使用 $redaction-mask-image-render 渲染最终脱敏图。")

# Text
OT = "$redaction-anonymize-text-flow"
add("非结构化", "纯文本TXT", OT, "—", "流程总调", "总调", "redaction-anonymize-text-flow", OT,
    "TXT 模块链", "请使用 $redaction-anonymize-text-flow 列出 TXT 脱敏模块顺序。")
add("非结构化", "纯文本TXT", OT, "2", "文本实体模块", "子步骤", "redaction-text-entity-module", "$redaction-text-entity-module",
    "NER", "使用 $redaction-text-entity-module 识别 TXT 中的敏感实体。")

# Batch
OB = "$redaction-anonymize-batch-flow"
add("非结构化", "批量", OB, "—", "流程总调", "总调", "redaction-anonymize-batch-flow", OB,
    "批量 API 流程", "请使用 $redaction-anonymize-batch-flow 说明批量任务 API 阶段（无 UI）。")

# Structured
OS = "$redaction-anonymize-structured-flow"
add("结构化", "表/库", OS, "—", "流程总调", "总调", "redaction-anonymize-structured-flow", OS,
    "结构化脱敏", "请使用 $redaction-anonymize-structured-flow 说明结构化数据脱敏阶段。")



def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Skill流程阶段表"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    orch_fill = PatternFill("solid", fgColor="E2EFDA")

    for col, name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(ROWS, 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row[5] == "总调":
                cell.fill = orch_fill

    widths = [18, 28, 32, 10, 22, 10, 32, 32, 36, 48]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT} ({len(ROWS)} rows)")


if __name__ == "__main__":
    main()
